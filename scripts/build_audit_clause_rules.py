from __future__ import annotations

import json
import re
from pathlib import Path

from docx import Document
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
AUDIT_ROOT = Path(r"Y:\IFC_V\公用區\115年感管查核")
SELF_ASSESSMENT = AUDIT_ROOT / "自評表" / "02-2.115年度醫院感染管制查核作業醫院自評表_上傳版V1(final)115.7.21更新.docx"
EVIDENCE_BOOK = AUDIT_ROOT / "自評表" / "115年感染管制查核條文佐證資料清單1150726(最新).xlsx"
OUT_FILE = BASE_DIR / "gas_line_bot" / "audit_clauses.json"


CLAUSE_GUIDANCE = {
    "1.1": {
        "aliases": ["感染管制會", "感管會", "委員會", "會議決議", "年度工作計畫"],
        "focus": "感染管制會組成須符合醫院業務，至少每3個月開會並留存院長批示之紀錄；年度工作計畫、跨部門分工、決議追蹤及前次缺失改善均須送會監督。",
        "questions": "感管會多久開一次？哪些部門參與？決議如何追蹤？年度計畫與前次查核缺失如何呈現執行成果？",
    },
    "1.2": {
        "aliases": ["感染管制單位", "感管中心", "感管人力", "感染管制師", "人力配置", "感管組織"],
        "focus": "感染管制單位須有明確組織定位、職責與實質資源；感染症醫師、感管護理及醫檢人力應按開放床數配置，具合格證書與前一年20學分，並有缺額代理及人才留任機制。",
        "questions": "目前開放床數與感管人力如何核算？證書與學分是否符合？缺額期間誰代理？院方提供哪些經費、空間與資訊支援？",
    },
    "1.3": {
        "aliases": ["感染管制手冊", "手冊更新", "感染管制教育", "教育", "教育訓練", "國際疫情", "疫情資訊", "新興傳染病教育"],
        "focus": "感染管制手冊應涵蓋標準與傳播途徑防護、侵入性裝置bundle、醫材與環境清消、廢棄物及特定區域管理，定期更新並依職別教育；同時持續蒐集國內外疫情與實證並傳達全院。",
        "questions": "手冊何時更新、由誰核定？同仁從哪裡取得最新版？教育如何依職別規劃？國際疫情如何蒐集、傳達並確認落實？",
    },
    "1.4": {
        "aliases": ["手部衛生", "洗手設備", "乾洗手", "濕洗手", "洗手稽核", "五時機"],
        "focus": "各醫療區須有適當洗手設備與point of care乾洗手液，訂有手部衛生程序；定期稽核五時機、正確率、設備完整性及乾洗手液用量，回饋缺失並追蹤改善。",
        "questions": "乾洗手與濕洗手何時使用？設備如何巡查？遵從率與正確率是多少？低落單位如何改善及再查核？",
    },
    "1.5": {
        "aliases": ["隔離措施", "隔離動線", "病人分流", "TOCC", "PPE穿脫", "大規模感染", "疫災應變", "負壓病室"],
        "focus": "依傳播途徑建立檢傷、分流、隔離、轉送與PPE流程，維持負壓與隔離病室管理；並備有大規模感染事件應變計畫、演練、物資與跨單位協調紀錄。",
        "questions": "疑似感染病人如何從入口分流？隔離床不足或負壓異常怎麼辦？PPE如何選擇與穿脫？大規模事件如何啟動、演練及檢討？",
    },
    "1.6": {
        "aliases": ["醫療照護相關感染", "HAI", "感染率", "THAS", "群聚", "群突發", "bundle成效"],
        "focus": "持續監測HAI、重要菌株、侵入性裝置及群聚事件，定期於相關會議分析趨勢並回饋單位；異常須調查、提出改善方案並追蹤具體成效。",
        "questions": "單位主要感染風險與趨勢為何？異常警示如何啟動？資料回饋到哪些會議？改善後如何證明有效？",
    },
    "1.7": {
        "aliases": ["安全注射", "注射安全", "單次使用", "藥品分裝", "多劑量藥瓶", "無菌操作"],
        "focus": "注射準備與執行須符合無菌技術及單次使用原則，安全處理針具、藥品與注射用品；對多劑量藥瓶、分裝區域及異常事件訂有管理與稽核。",
        "questions": "注射藥品在哪裡準備？單次使用用品能否共用？多劑量藥瓶如何管理？現場如何稽核安全注射？",
    },
    "2.1": {
        "aliases": ["衛材清潔", "器械消毒", "滅菌", "內視鏡再處理", "高層次消毒", "醫療儀器清消", "AER"],
        "focus": "依Spaulding分類及廠商說明完成器械、機器與內視鏡的清洗、消毒或滅菌；流程須包含人員訓練、清污分流、設備維護、消毒劑監測、乾燥儲存及完整追溯。",
        "questions": "哪些物品需滅菌或高層次消毒？清洗品質如何確認？內視鏡如何測漏、乾燥與儲存？異常批次如何停用、召回及改善？",
    },
    "2.2": {
        "aliases": ["供應室", "CSSD", "清潔區污染區", "無菌物品", "滅菌物品配發", "供應室動線"],
        "focus": "供應室須配置適當且受訓人力，污染區、清潔區及無菌區應有明確實體區隔與單向動線；落實環境、包裝、儲存、運送及設備品質監測。",
        "questions": "清污與無菌區如何區隔？人員與物品動線是否交叉？滅菌物品如何儲存與配發？環境及設備異常如何處理？",
    },
    "2.3": {
        "aliases": ["環境清潔", "環境消毒", "終期清潔", "清潔人員", "消毒劑", "高頻表面", "病室清消"],
        "focus": "依區域與感染風險訂定清潔頻率、工具分區、消毒劑濃度與接觸時間；管理高頻表面、共用設備及終期清潔，並以教育、紀錄、稽核與改善確保品質。",
        "questions": "清潔頻率與責任如何訂定？消毒劑如何泡製與標示？隔離病人轉出後如何執行及確認終期清潔？缺失如何追蹤？",
    },
    "3.1": {
        "aliases": ["抗生素管理計畫", "抗菌藥物管理", "ASP", "抗生素管理小組", "抗生素領導"],
        "focus": "由院級領導支持跨專業抗生素管理計畫，明訂權責、專責人力、資訊及資源，定期於委員會檢討目標、策略與執行成果。",
        "questions": "誰負責領導抗生素管理？跨專業成員如何分工？年度目標、資源與成果如何向院方報告？",
    },
    "3.2": {
        "aliases": ["抗生素使用監測", "抗生素用量", "抗生素處方", "管制性抗生素", "手術預防性抗生素", "de-escalation"],
        "focus": "建立抗生素用量、處方適當性、培養採檢、管制藥審核、去升階或停藥及手術預防性抗生素監測，回饋處方者並追蹤改善。",
        "questions": "如何監測用量與適當性？何時重新評估或停藥？異常科別如何回饋？手術預防性抗生素指標如何追蹤？",
    },
    "3.3": {
        "aliases": ["抗藥性監測", "抗藥菌管理", "MDRO", "VRE", "CRE", "CRAB", "CRPA", "MRSA", "抗生素抗藥性"],
        "focus": "持續監測抗藥性微生物與趨勢，檢驗結果應及時通知並連結隔離、床位、器材清消、轉送及終期清潔；異常與群聚須調查並追蹤改善。",
        "questions": "重要抗藥菌如何定義與警示？陽性後如何通知及隔離？再入院如何辨識？群聚或趨勢上升時如何處理？",
    },
    "4.1": {
        "aliases": ["傳染病通報", "法定傳染病", "通報機制", "衛生局聯繫", "疾病通報", "疫情通報"],
        "focus": "建立疑似法定傳染病的辨識、通報、採檢、隔離及聯繫機制，指定專責人員確認通報時效、資料完整性及與衛生主管機關之後續協調。",
        "questions": "誰負責通報與複核？夜間假日如何處理？如何確認沒有逾時或漏報？衛生局要求補件時如何追蹤？",
    },
    "4.2": {
        "aliases": ["痰液耐酸性塗片陽性", "耐酸性塗片陽性", "AFB陽性", "疑似肺結核", "結核隔離"],
        "focus": "痰液耐酸性塗片陽性或疑似具傳染性結核個案應立即採空氣隔離、安排適當病室與轉送防護，完成通報、檢體追蹤、環境及暴露風險管理。",
        "questions": "AFB陽性後如何安置與通報？外出檢查如何防護？何時可解除空氣隔離？接觸者如何啟動追蹤？",
    },
    "4.3": {
        "aliases": ["結核個案管理", "結核個管", "都治", "結核衛教", "TB個管"],
        "focus": "指定人員負責結核病個案管理、通報、治療銜接、都治與衛教，持續追蹤服藥、檢驗、轉歸及失聯個案，並與公衛單位協調。",
        "questions": "個案由誰管理？出院後如何銜接都治？服藥與檢驗如何追蹤？失聯或轉院個案如何處理？",
    },
    "4.4": {
        "aliases": ["結核診治", "TB診治", "結核檢驗", "結核治療", "MDRTB", "抗藥性結核"],
        "focus": "具備結核病診斷、檢驗、治療與跨科照護機制，依規範完成痰檢、培養、藥敏及個案分類；疑似抗藥性或特殊個案應及時轉介與專案管理。",
        "questions": "疑似結核如何安排檢驗與治療？檢驗結果如何追蹤？抗藥性或複雜個案由誰會診及轉介？",
    },
    "4.5": {
        "aliases": ["結核接觸者", "院內接觸者", "TB接觸者", "結核暴露", "接觸者追蹤"],
        "focus": "住院確診具傳染性結核時，須依可傳染期、病室動線與防護情形完整匡列院內接觸者，完成通知、檢查、列管及追蹤結果。",
        "questions": "如何界定可傳染期與接觸者？工作人員及病人如何通知與檢查？未完成追蹤者如何管理？",
    },
    "4.6": {
        "aliases": ["防護裝備儲備", "防疫物資", "PPE庫存", "N95庫存", "口罩儲備", "物資效期", "安全庫存"],
        "focus": "依主管機關規定及風險評估維持口罩、N95、隔離衣、手套等防護裝備安全庫存，管理合格證明、效期、輪替、盤點、領用及緊急調度。",
        "questions": "安全庫存量如何設定？物資是否符合規格？效期與先進先出如何管理？疫情大量需求時如何調度？",
    },
    "4.7": {
        "aliases": ["透析肝炎", "透析室B肝", "透析室C肝", "透析分區", "透析機消毒", "透析肝炎監測"],
        "focus": "透析工作人員及病人須定期監測B、C型肝炎；B肝病人分區分床機、C肝病人集中照護，班次間完成透析機與照護區清消，陽轉者須追蹤及依法通報。",
        "questions": "員工與病人多久檢查一次？B肝如何分區分機？C肝如何集中照護？每位病人後的機器與環境如何消毒？陽轉如何處理？",
    },
    "5.1": {
        "aliases": ["疫苗政策", "預防接種", "員工疫苗", "健康監測", "胸部X光", "胸部X光檢查", "MMR", "流感疫苗", "新冠疫苗"],
        "focus": "訂定全院工作人員疫苗、健康監測及胸部X光保護計畫，涵蓋適用的外包、實習與志工；追蹤接種率、麻疹免疫力、健康異常、胸部X光達成率及暴露後處置。",
        "questions": "流感、MMR、COVID-19及B肝疫苗對象與達成率為何？未接種或抗體陰性如何追蹤？健康異常與胸部X光結果如何管理？",
    },
    "5.2": {
        "aliases": ["針扎", "尖銳物扎傷", "血液體液暴露", "切傷", "職業暴露", "HIV PEP", "暴露後處理"],
        "focus": "訂有尖銳物與血液體液暴露預防、立即處理、通報、風險評估、檢驗、預防用藥、心理支持及追蹤流程；事件須統計分析、回饋並改善。",
        "questions": "暴露後第一步做什麼？何時啟動HIV PEP？通報、檢驗與追蹤在哪裡完成？事件原因及改善如何回饋？",
    },
}


def compact(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def load_titles() -> dict[str, str]:
    doc = Document(str(SELF_ASSESSMENT))
    titles: dict[str, str] = {}
    for table in doc.tables:
        for row in table.rows:
            values = [compact(cell.text) for cell in row.cells]
            if len(values) >= 3 and re.fullmatch(r"[1-5]\.[1-7]", values[1]):
                titles[values[1]] = values[2]
    return titles


def load_evidence() -> dict[str, list[dict[str, str]]]:
    workbook = load_workbook(str(EVIDENCE_BOOK), read_only=True, data_only=True)
    sheet = workbook["115年佐證資料清單"]
    evidence: dict[str, list[dict[str, str]]] = {key: [] for key in CLAUSE_GUIDANCE}
    seen_names: dict[str, set[str]] = {key: set() for key in CLAUSE_GUIDANCE}
    for row in sheet.iter_rows(values_only=True):
        values = [compact(value) for value in row]
        if len(values) < 4:
            continue
        item = values[1]
        match = re.match(r"^([1-5]\.[1-7])-(?:符|優)\d+", item)
        if not match or match.group(1) not in evidence:
            continue
        name, url = values[2], values[3]
        if not name or name in seen_names[match.group(1)]:
            continue
        normalized_url = re.sub(r"^http://km\.ntuh\.gov\.tw/", "https://km.ntuh.gov.tw/", url, flags=re.I)
        km_url = normalized_url if re.match(r"^https://km\.ntuh\.gov\.tw/", normalized_url, flags=re.I) else ""
        evidence[match.group(1)].append({"name": name, "url": km_url})
        seen_names[match.group(1)].add(name)
    for clause_id in evidence:
        evidence[clause_id].sort(key=lambda item: (0 if item["url"] else 1, item["name"]))
    return evidence


def main() -> None:
    titles = load_titles()
    evidence = load_evidence()
    missing_titles = sorted(set(CLAUSE_GUIDANCE) - set(titles))
    if missing_titles:
        raise RuntimeError(f"Missing official clause titles: {missing_titles}")

    clauses = []
    for clause_id in sorted(CLAUSE_GUIDANCE, key=lambda value: tuple(map(int, value.split(".")))):
        guide = CLAUSE_GUIDANCE[clause_id]
        clauses.append({
            "id": clause_id,
            "title": titles[clause_id],
            "aliases": guide["aliases"],
            "focus": guide["focus"],
            "questions": guide["questions"],
            "evidence": evidence.get(clause_id, []),
        })

    payload = {
        "version": "115.7.26",
        "official_clause_count": len(clauses),
        "source_self_assessment": SELF_ASSESSMENT.name,
        "source_evidence": EVIDENCE_BOOK.name,
        "clauses": clauses,
    }
    OUT_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "output": str(OUT_FILE),
        "clauses": len(clauses),
        "evidence_items": sum(len(item["evidence"]) for item in clauses),
        "clauses_without_evidence": [item["id"] for item in clauses if not item["evidence"]],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
